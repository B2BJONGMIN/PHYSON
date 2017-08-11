from openpyxl import Workbook
from openpyxl import load_workbook
import os
from openpyxl.styles import Font, Alignment
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill, Color
# import requests
# import lxml
# import re
# import openpyxl
# import html5lib
import requests
from bs4 import BeautifulSoup



# 엑셀파일에 검색한 세미나 Inert
def Insert_Into_Excel_File(Seminar_List):
    wb = load_workbook("ITseminar.xlsx")
    worksheet = wb.active

    box = Border(
        left=Side(border_style="thin", color='FF000000'),
        right=Side(border_style="thin", color='FF000000'),
        top=Side(border_style="thin", color='FF000000'),
        bottom=Side(border_style="thin", color='FF000000'),
    )

    for i in range(len(Seminar_List)):
        row = int(worksheet.max_row)
        worksheet.row_dimensions[row + 1].height = 66

        if (i % 5 == 0):
            worksheet.cell(row=row + 1, column=1).value = row - 1
            worksheet.cell(row=row + 1, column=2).value = "제목 : " + Seminar_List[i]
            worksheet.cell(row=row + 1, column=1).alignment = Alignment(horizontal='center', vertical='center')

        elif (i % 5 == 1):
            worksheet.cell(row=row, column=2).value = str(worksheet.cell(row=row, column=2).value) + "\n일시 : " + \
                                                      Seminar_List[i]

        elif (i % 5 == 2):
            worksheet.cell(row=row, column=2).value = str(worksheet.cell(row=row, column=2).value) + "\n장소 : " + \
                                                      Seminar_List[i]
            worksheet.cell(row=row, column=2).alignment = Alignment(horizontal='left', vertical='center',
                                                                    wrap_text=True)

        elif (i % 5 == 3):
            worksheet.cell(row=row, column=3).value = Seminar_List[i]
            worksheet.cell(row=row, column=3).alignment = Alignment(horizontal='center', vertical='center',
                                                                    wrap_text=True)

        elif (i % 5 == 4):
            worksheet.cell(row=row, column=4).value = Seminar_List[i]
            worksheet.cell(row=row, column=4).alignment = Alignment(horizontal='center', vertical='center',
                                                                    wrap_text=True)

    for row_num in range(1, int(worksheet.max_row) + 1):
        for col_num in range(1, 5):
            worksheet.cell(row=row_num, column=col_num).border = box

    wb.save("ITseminar.xlsx")  # 엑셀 파일 저장


# http://www.sek.co.kr/schedule_all.php? 크롤링
def CrawlingFromSek():
    Seminar_Info = []
    loop_count = 0

    while (True):

        if (loop_count == 5):
            break
        # 세미나 url 크롤링 Start
        #  └ url : 세미나상세정보 url
        Domain = 'http://www.sek.co.kr'
        res = requests.get('http://www.sek.co.kr/schedule_all.php?')
        #soup = BeautifulSoup(res.text, 'lxml')
       # soup= BeautifulSoup(res.text,'html5lib')
        soup = BeautifulSoup(res.text, 'html.parser')

        tag = soup.find(attrs={'id': 'board'}).find_all(attrs={'class': 'float_clear schedule_list schedule_pad'})[
            loop_count].find_all('a')[0]
        url = Domain + str(tag).split('"')[1]
        # 세미나 url 크롤링 End


        # 해당 세미나 url타고 들어가 크롤링 Start
        #  ├SeminarName : 세미나 이름
        #  ├SeminarDate : 세미나 날짜
        #  ├SeminarLoc  : 세미나 위치
        #  └SeminarCost : 세미나 비용
        res = requests.get(url)
       #soup = BeautifulSoup(res.text, 'lxml')
        soup = BeautifulSoup(res.text, 'html.parser')

        tag = soup.find(attrs={'id': 'view_info'}).find('table')

        SeminarName = tag.find_all('tr')[1].find_all('td')[0].text.replace("\t", "").replace("\r", "").replace("\n", "")
        SeminarDate = tag.find_all('tr')[2].find_all('td')[0].text.replace("\t", "").replace("\r", "").replace("\n", "")
        SeminarLoc = tag.find_all('tr')[3].find_all('td')[0].text.replace("\t", "").replace("\r", "").replace("\n", "")
        SeminarCost = tag.find_all('tr')[6].find_all('td')[0].text.replace("\t", "").replace("\r", "").replace("\n", "")
        # 해당 세미나 url타고 들어가 크롤링 End

        # 해당 세미나 리스트에 append Start
        Seminar_Info.append(SeminarName)
        Seminar_Info.append(SeminarDate)
        Seminar_Info.append(SeminarLoc)
        Seminar_Info.append(SeminarCost)
        Seminar_Info.append("")
        # 해당 세미나 리스트에 append End

        loop_count = loop_count + 1

    return Seminar_Info


# Make ExcelFile
def excel_file_create():
    # Excel 파일 생성
    #   ├ worksheet       = 현재 열려 있는 Sheet
    #   ├ worksheet.title = 시트의 이름
    wb = Workbook()
    worksheet = wb.active
    worksheet.title = "Seminar"

    worksheet.merge_cells("A1:D1")
    worksheet["A1"] = "ICT 세미나/컨퍼런스"  # 셀 A1에 입력
    worksheet["A2"] = "NO"
    worksheet["B2"] = "주요내용"
    worksheet["C2"] = "비용"
    worksheet["D2"] = "신청"

    # 테두리 값 설정
    box = Border(
        left=Side(border_style="thin", color='FF000000'),
        right=Side(border_style="thin", color='FF000000'),
        top=Side(border_style="thin", color='FF000000'),
        bottom=Side(border_style="thin", color='FF000000'),
    )

    # 테두리 범위 설정
    for row_num in range(1, 3):
        for col_num in range(1, 5):
            worksheet.cell(row=row_num, column=col_num).border = box

    ca = worksheet["A1"]
    ca.font = Font(name='맑은 고딕', size=18, bold=True)
    ca.alignment = Alignment(horizontal='center', vertical='center')

    for col_num in range(1, 5):
        worksheet.cell(row=2, column=col_num).font = Font(name='맑은 고딕', size=14, bold=True)
        worksheet.cell(row=2, column=col_num).alignment = Alignment(horizontal='center', vertical='center')

    # 셀 크기
    worksheet.column_dimensions["A"].width = 5
    worksheet.column_dimensions["B"].width = 64.88
    worksheet.column_dimensions["C"].width = 20
    worksheet.column_dimensions["D"].width = 14

    wb.save("ITseminar.xlsx")  # 엑셀 파일 저장


# 파일 존재 여부 확인
def file_exist_detector():
    for root, dirs, files in os.walk('./'):
        for file in files:
            if (file == 'ITseminar.xlsx'):
                return True
    return False


def main():
    # 엑셀 템플릿 없으면 생성
    if (not file_exist_detector()):
        excel_file_create()

    # Sek.co.kr 세미나 크롤링
    Seminar_List = CrawlingFromSek()
    Insert_Into_Excel_File(Seminar_List)

    return 0


if __name__ == '__main__':
    main()

